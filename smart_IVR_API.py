import pandas as pd
import openpyxl
from openpyxl import load_workbook
import matplotlib.pyplot as plt
from fastapi import FastAPI, UploadFile, File, HTTPException, Request
from fastapi.responses import StreamingResponse, JSONResponse, FileResponse
import zipfile
import io
import tempfile
import os
import traceback
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def normalize_phone(series):
    s = (
        series.astype(str)
        .str.replace(r"\.0$", "", regex=True)
        .str.replace(r"\D", "", regex=True)   # remove everything not digit
        .str.replace(r"^389", "", regex=True) # remove country code if exists
        .str.lstrip("0")                      # remove leading zero
    )

    s = s[s.str.len() >= 8]  # keep valid numbers only
    return "+389" + s


class ExcelWorker:
    def __init__(self):
        self.df_povici = None
        self.df_precki = None

    def read_excel_create_dfs(
        self,
        m1p, m2p, m3p,
        m1pr, m2pr, m3pr,
        temp_dir
    ):
        logger.info("Reading Excel files...")

        self.df_povici = pd.concat([
            pd.read_excel(m1p),
            pd.read_excel(m2p),
            pd.read_excel(m3p)
        ], ignore_index=True)

        self.df_precki = pd.concat([
            pd.read_excel(m1pr),
            pd.read_excel(m2pr),
            pd.read_excel(m3pr)
        ], ignore_index=True)

        def clean_col(df, col):
            if col in df.columns:
                df[col] = (
                    df[col]
                    .astype(str)
                    .str.replace("\xa0", " ", regex=False)
                    .str.strip()
                    .str.lower()
                )
            return df

        self.df_precki = clean_col(self.df_precki, "Тип на пречката")
        self.df_precki = clean_col(self.df_precki, "Status nalog")

        self.df_precki["Класификација"] = (
            self.df_precki["Класификација"]
            .astype(str)
            .str.replace("\xa0", " ", regex=False)
            .str.strip()
            .str.upper()
        )

        self.df_precki["Категорија"] = (
            self.df_precki["Категорија"]
            .astype(str)
            .str.replace("\xa0", " ", regex=False)
            .str.strip()
        )

        logger.info(f"Before filter: {len(self.df_precki)} rows")

        self.df_precki = self.df_precki[
            (self.df_precki["Status nalog"] != "откажан") &
            (self.df_precki["Тип на пречката"] != "network facing") &
            (~self.df_precki["Класификација"].isin([
                "WHOLESALE", "MOBILE  POSTPAID", "MOBILE PREPAID"
            ])) &
            (self.df_precki["Категорија"] == "Физичко лице")
        ]

        logger.info(f"After filter: {len(self.df_precki)} rows")

        self.df_precki["Контакт"] = normalize_phone(self.df_precki["Контакт"])

        # drop rows that became NaN after filtering
        self.df_precki = self.df_precki.dropna(subset=["Контакт"])

        # debug export
        self.df_precki.to_excel(os.path.join(temp_dir, "prechki_combined.xlsx"), index=False)

        # filter povici
        self.df_povici = self.df_povici[
            self.df_povici["Direction"] == "Inbound"
        ]

        return self.df_povici, self.df_precki

    def remove_FTTH_ready(self, ftth_path):
        logger.info("Removing FTTH Ready...")

        ffth_ready = pd.read_excel(ftth_path)

        ffth_ready["CINUMS"] = ffth_ready["CINUMS"].astype(str).str.strip()
        self.df_precki["LineID"] = self.df_precki["LineID"].astype(str).str.strip()

        self.df_precki = self.df_precki[
            ~self.df_precki["LineID"].isin(ffth_ready["CINUMS"])
        ]

    def mark_slabi_linii(self, slabi_path):
        logger.info("Marking weak lines...")

        slabi_linii = pd.read_excel(slabi_path, header=8)
        slabi_linii["Row Labels"] = slabi_linii["Row Labels"].astype(str).str.strip()

        self.df_precki["LineID"] = self.df_precki["LineID"].astype(str).str.strip()

        self.df_precki["слаба линија"] = self.df_precki["LineID"].apply(
            lambda x: "Да" if x in slabi_linii["Row Labels"].values else "Не"
        )

    def create_report(self, output_path):
        logger.info("Creating report...")

        counter = Counter(self.df_povici, self.df_precki)

        df_precki = counter.count_precki()
        df_povici = counter.count_povici()

        merged = (
            df_precki
            .merge(df_povici, on="телефонски број", how="outer")
            .fillna(0)
        )

        merged["отворени пречки"] = merged["отворени пречки"].astype(int)
        merged["број на повици во контакт центар"] = merged["број на повици во контакт центар"].astype(int)

        merged = merged[
            (merged["отворени пречки"] >= 2) &
            (merged["број на повици во контакт центар"] >= 2)
        ]

        merged = merged[
            ["LineID", "телефонски број", "отворени пречки",
             "број на повици во контакт центар", "слаба линија"]
        ]

        merged.to_excel(output_path, index=False)
        return merged


class Counter:
    def __init__(self, df_povici, df_precki):
        self.df_povici = df_povici
        self.df_precki = df_precki

    def count_precki(self):
        df = self.df_precki.copy()

        df["телефонски број"] = df["Контакт"]

        result = (
            df.groupby("телефонски број")
            .agg({
                "LineID": "first",
                "Контакт": "size",
                "слаба линија": "first"
            })
            .reset_index()
            .rename(columns={"Контакт": "отворени пречки"})
        )

        return result

    def count_povici(self):
        col = "TBP_ANI (Case) (Old Value)"

        if col not in self.df_povici.columns:
            raise KeyError(f"Missing column: {col}")

        df = normalize_phone(self.df_povici[col])

        df = df.value_counts().reset_index()
        df.columns = ["телефонски број", "број на повици во контакт центар"]

        return df


class Histogram:
    def __init__(self, df):
        self.df = df

    def scatter_plot(self, path):
        plt.figure(figsize=(10, 6))
        plt.scatter(
            self.df["отворени пречки"],
            self.df["број на повици во контакт центар"]
        )
        plt.xlabel("Отворени пречки")
        plt.ylabel("Број на повици во контакт центар")
        plt.grid()
        plt.tight_layout()
        plt.savefig(path)
        plt.close()


def styling(report_path, cinums_path):
    wb = load_workbook(report_path)
    sheet = wb.active

    widths = [15, 30, 30, 40, 20]
    for i, w in enumerate(widths, start=1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    wb.save(report_path)

    wb = load_workbook(cinums_path)
    sheet = wb.active
    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].width = 30
    wb.save(cinums_path)


app = FastAPI()


@app.exception_handler(Exception)
async def global_exception_handler(request: Request, exc: Exception):
    logger.error(traceback.format_exc())
    return JSONResponse(status_code=500, content={"error": str(exc)})


@app.get("/")
def home():
    return FileResponse("index.html")


@app.post("/process")
async def process_files(
    month1_povici: UploadFile = File(...),
    month2_povici: UploadFile = File(...),
    month3_povici: UploadFile = File(...),
    month1_prechki: UploadFile = File(...),
    month2_prechki: UploadFile = File(...),
    month3_prechki: UploadFile = File(...),
    ftth_ready: UploadFile = File(...),
    slabi_linii: UploadFile = File(...),
):
    temp_dir = tempfile.mkdtemp()

    try:
        paths = {}
        for name, file in {
            "m1p": month1_povici,
            "m2p": month2_povici,
            "m3p": month3_povici,
            "m1pr": month1_prechki,
            "m2pr": month2_prechki,
            "m3pr": month3_prechki,
            "ftth": ftth_ready,
            "slabi": slabi_linii,
        }.items():
            path = os.path.join(temp_dir, f"{name}.xlsx")
            with open(path, "wb") as f:
                f.write(await file.read())
            paths[name] = path

        worker = ExcelWorker()
        worker.read_excel_create_dfs(
            paths["m1p"], paths["m2p"], paths["m3p"],
            paths["m1pr"], paths["m2pr"], paths["m3pr"],
            temp_dir
        )

        worker.remove_FTTH_ready(paths["ftth"])
        worker.mark_slabi_linii(paths["slabi"])

        report_path = os.path.join(temp_dir, "report.xlsx")
        final_df = worker.create_report(report_path)

        hist_path = os.path.join(temp_dir, "histogram.png")
        Histogram(final_df).scatter_plot(hist_path)

        cinums_path = os.path.join(temp_dir, "CINUMS.xlsx")

        filtered = final_df[final_df["слаба линија"] != "Да"]

        pd.DataFrame({
            "CINUMS": filtered["LineID"],
            "Reason Code": "RepeatedTCCAgent"
        }).to_excel(cinums_path, index=False)

        styling(report_path, cinums_path)

        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w") as z:
            z.write(report_path, "report.xlsx")
            z.write(cinums_path, "CINUMS.xlsx")
            z.write(hist_path, "histogram.png")

        zip_buffer.seek(0)

        return StreamingResponse(
            zip_buffer,
            media_type="application/zip",
            headers={"Content-Disposition": "attachment; filename=results.zip"}
        )

    except Exception as e:
        logger.error(traceback.format_exc())
        raise HTTPException(500, str(e))